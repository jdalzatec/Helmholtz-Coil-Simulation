import sys

is_frozen = getattr(sys, 'frozen', False)
frozen_temp_path = getattr(sys, '_MEIPASS', '')

import os

# This is needed to find resources when using pyinstaller
if is_frozen:
    basedir = frozen_temp_path
else:
    basedir = os.path.dirname(os.path.abspath(__file__))
resource_dir = os.path.join(basedir, 'resources')




import gi
gi.require_version("Gtk", "3.0")
from gi.repository import Gtk

from matplotlib import pyplot
from matplotlib.figure import Figure
from matplotlib.backends.backend_gtk3agg import FigureCanvasGTK3Agg as FigureCanvas
# from matplotlib.backends.backend_gtk3cairo import FigureCanvasGTK3Cairo as FigureCanvas
from matplotlib.backends.backend_gtk3 import NavigationToolbar2GTK3 as NavigationToolbar

import numpy
from PlotWindow import PlotBox
from ZoomWindow import ZoomWindow
from HomogeneityWindow import HomogeneityWindow
from coil import Coil, CreateCoil
from CoilListRow import CoilListRow
from About import AboutWindow

import openpyxl


class Results():
    def __init__(self, parent, simulation):
        self.parent = parent
        
        self.builder = Gtk.Builder()
        
        self.builder.add_from_file(resource_dir + "/results.glade")
        self.window = self.builder.get_object("wndResults")
        self.boxPlot = self.builder.get_object("boxPlot")
        self.btnBack = self.builder.get_object("btnBack")
        self.btnZoom = self.builder.get_object("btnZoom")
        self.btnHomogeneity = self.builder.get_object("btnHomogeneity")
        self.statBar = self.builder.get_object("statBar")
        self.menuColorMap = self.builder.get_object("menuColorMap")
        self.txtInputParameters = self.builder.get_object("txtInputParameters")
        self.txtElectircalParameters = self.builder.get_object("txtElectircalParameters")
        self.btnSaveAs = self.builder.get_object("btnSaveAs")
        self.btnOpen = self.builder.get_object("btnOpen")
        self.btnQuit = self.builder.get_object("btnQuit")
        self.btnAbout = self.builder.get_object("btnAbout")

        self.simulation = simulation

        self.colormap = "jet"

        self.window.set_transient_for(self.parent.window)

        self.window.connect("destroy", self.quit)
        self.btnBack.connect("clicked", self.on_back)
        self.btnZoom.connect("clicked", self.on_zoom)
        self.btnHomogeneity.connect("clicked", self.on_homogeneity)
        self.btnSaveAs.connect("activate", self.on_export)
        self.btnOpen.connect("activate", self.on_import)
        self.btnQuit.connect("activate", Gtk.main_quit)
        self.btnAbout.connect("activate", lambda _: AboutWindow(self.window))


        # Get a list of the colormaps in matplotlib.  Ignore the ones that end with
        # '_r' because these are simply reversed versions of ones that don't end
        # with '_r'
        maps = sorted(m for m in pyplot.cm.datad if not m.endswith("_r"))

        firstitem = Gtk.RadioMenuItem(self.colormap)
        firstitem.set_active(True)
        firstitem.connect('activate', self.on_color_bar_menu, self.colormap)
        self.menuColorMap.append(firstitem)
        for name in maps:
            if name != self.colormap:
                item = Gtk.RadioMenuItem.new_with_label([firstitem], name)
                item.set_active(False)
                item.connect('activate', self.on_color_bar_menu, name)
                self.menuColorMap.append(item)

        self.load_simulation()
        
        self.window.maximize()
        self.window.show_all()



    def load_simulation(self):
        z_arr = [coil.pos_z for coil in self.simulation.coils]
        radius_arr = [coil.radius for coil in self.simulation.coils]
        
        xlims = (min(z_arr), max(z_arr))
        ylims = (-max(radius_arr), max(radius_arr))

        for ch in self.boxPlot:
            self.boxPlot.remove(ch)

        self.plot = PlotBox(self, self.simulation,
            self.colormap, self.statBar, xlims, ylims)
        self.boxPlot.pack_start(self.plot.boxPlot, True, True, 0)
        self.populate_input_parameters()
        self.populate_electrical_parameters()

        self.window.show_all()




    def on_zoom(self, widget):
        self.plot.clear_rectangle()
        zoom = ZoomWindow(self, self.simulation, self.colormap)

    def on_homogeneity(self, widget):
        homogeneity = HomogeneityWindow(self, self.simulation, self.colormap)

    def populate_input_parameters(self):
        text = "\n"
        text += "\t{}\t\t=\t\t{}\n".format("Min. z [m]", str(self.simulation.z_min))
        text += "\t{}\t\t=\t\t{}\n".format("Max. z [m]", str(self.simulation.z_max))
        text += "\t{}\t\t\t=\t\t{}\n".format("Points z", str(self.simulation.z_points))
        text += "\t{}\t\t=\t\t{}\n".format("Min. y [m]", str(self.simulation.y_min))
        text += "\t{}\t\t=\t\t{}\n".format("Max. y [m]", str(self.simulation.y_max))
        text += "\t{}\t\t\t=\t\t{}\n".format("Points y", str(self.simulation.y_points))

        text += "\n"

        text += "\t{}\t\t{}\t\t{}\t\t{}\n".format(
            "Radius [m]", "Turns", "Current [A]", "Position [m]")

        for coil in self.simulation.coils:
            text += "\t{:.5f}\t\t\t{:d}\t\t\t{:.5f}\t\t\t{:.5f}\n".format(
                coil.radius, coil.num_turns, coil.I, coil.pos_z)

        self.txtInputParameters.get_buffer().set_text(text)


    def populate_electrical_parameters(self):
        gauge, diameter, section, resist, Inominal = numpy.loadtxt(resource_dir + "/awg.dat", unpack=True)
        Imax = max([abs(coil.I) for coil in self.simulation.coils])
        index = numpy.argmin(Inominal > Imax) - 1
        gauge = gauge[index]
        diameter = diameter[index]
        section = section[index]
        resist = resist[index]
        Inominal = Inominal[index]

        length = sum([2*numpy.pi*coil.radius*coil.num_turns for coil in self.simulation.coils]) * 1.05
        
        self.electrical_values = {
            "AWG Gauge": int(gauge),
            "Wire diameter [mm]": diameter,
            "Wire cross sectional area [mm2]": section,
            "Nominal current [A]": Inominal,
            "Maximum current [A]": Inominal * 1.1,
            "Total wire length [m]": length,
            "Wire resistance [Ohm]": resist * length / 1000,
            }


        text = "\n"
        text += "\t{}\t\t\t\t\t\t=\t\t{:d}\n".format("AWG Gauge", self.electrical_values["AWG Gauge"])
        text += "\t{}\t\t\t\t=\t\t{:.5f}\n".format("Wire diameter [mm]", self.electrical_values["Wire diameter [mm]"])
        text += "\t{}\t=\t\t{:.5f}\n".format("Wire cross sectional area [mm²]", self.electrical_values["Wire cross sectional area [mm2]"])
        text += "\t{}\t\t\t\t=\t\t{:.5f}\n".format("Nominal current [A]", self.electrical_values["Nominal current [A]"])
        text += "\t{}\t\t\t\t=\t\t{:.5f}\n".format("Maximum current [A]", self.electrical_values["Maximum current [A]"])
        text += "\t{}\t\t\t\t=\t\t{:.5f}\n".format("Total wire length [m]", self.electrical_values["Total wire length [m]"])
        text += "\t{}\t\t\t=\t\t{:.5f}\n".format("Wire resistance [Ω]", self.electrical_values["Wire resistance [Ohm]"])

        self.txtElectircalParameters.get_buffer().set_text(text)



    def on_color_bar_menu(self, widget, name):
        self.colormap = name
        self.plot.update_plot(name)


    def quit(self, widget):
        if not self.parent.window.get_visible():
            Gtk.main_quit()
        else:
            self.window.close()


    def on_back(self, widget):
        coil_rows = []
        for coil in self.simulation.coils:
            coil_row = CoilListRow()
            coil_row.set_values(
                radius=coil.radius,
                turns=coil.num_turns,
                current=coil.I, position=coil.pos_z)
            coil_rows.append(coil_row)
        self.parent.listBox.update(coil_rows)
        self.parent.window.show()
        self.window.close()


    def on_export(self, widget):
        dialog = Gtk.FileChooserDialog("Please choose a file", self.window,
            Gtk.FileChooserAction.SAVE,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_SAVE, Gtk.ResponseType.OK))

        filters = Gtk.FileFilter()
        filters.set_name("Excel files")
        filters.add_pattern("*.*.csv")
        filters.add_pattern("*.*.CSV")
        filters.add_pattern("*.xlsx")
        filters.add_pattern("*.XLSX")
        dialog.add_filter(filters)

        response = dialog.run()
        
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()
            # print("Open clicked")
            # print("File selected: " + filename)
            if "." not in filename:
                filename += ".xlsx"


            wb = openpyxl.Workbook()
            wb.remove_sheet(wb.active)

            wInput = wb.create_sheet('Simulation parameters')
            wCoils = wb.create_sheet('Input parameters')
            wElectrical = wb.create_sheet('Electrical parameters')
            wBy = wb.create_sheet('B y')
            wBz = wb.create_sheet('B z')
            wBnorm = wb.create_sheet('B norm')
            title_style = openpyxl.styles.Font(bold=True) 

            wInput.cell(row=1 + 0, column=1 + 0).value = "Min. z [m]"
            wInput.cell(row=1 + 0, column=1 + 0).font = title_style
            wInput.cell(row=1 + 0, column=1 + 1).value = self.simulation.z_min
            
            wInput.cell(row=1 + 1, column=1 + 0).value = "Max. z [m]"
            wInput.cell(row=1 + 1, column=1 + 0).font = title_style
            wInput.cell(row=1 + 1, column=1 + 1).value = self.simulation.z_max
            
            wInput.cell(row=1 + 2, column=1 + 0).value = "Points z"
            wInput.cell(row=1 + 2, column=1 + 0).font = title_style
            wInput.cell(row=1 + 2, column=1 + 1).value = self.simulation.z_points - 1

            wInput.cell(row=1 + 3, column=1 + 0).value = "Min. y [m]"
            wInput.cell(row=1 + 3, column=1 + 0).font = title_style
            wInput.cell(row=1 + 3, column=1 + 1).value = self.simulation.y_min

            wInput.cell(row=1 + 4, column=1 + 0).value = "Max. y [m]"
            wInput.cell(row=1 + 4, column=1 + 0).font = title_style
            wInput.cell(row=1 + 4, column=1 + 1).value = self.simulation.y_max

            wInput.cell(row=1 + 5, column=1 + 0).value = "Points y"
            wInput.cell(row=1 + 5, column=1 + 0).font = title_style
            wInput.cell(row=1 + 5, column=1 + 1).value = self.simulation.y_points - 1

            wCoils.cell(row=1 + 0, column=1 + 0).value = "Radius [m]"
            wCoils.cell(row=1 + 0, column=1 + 0).font = title_style

            wCoils.cell(row=1 + 0, column=1 + 1).value = "Num. turns"
            wCoils.cell(row=1 + 0, column=1 + 1).font = title_style

            wCoils.cell(row=1 + 0, column=1 + 2).value = "Current [A]"
            wCoils.cell(row=1 + 0, column=1 + 2).font = title_style

            wCoils.cell(row=1 + 0, column=1 + 3).value = "Pos. Z [m]"
            wCoils.cell(row=1 + 0, column=1 + 3).font = title_style


            wElectrical.cell(row=1 + 0, column=1 + 0).value = "AWG Gauge"
            wElectrical.cell(row=1 + 0, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 0, column=1 + 1).value = self.electrical_values["AWG Gauge"]
            wElectrical.cell(row=1 + 1, column=1 + 0).value = "Wire diameter [mm]"
            wElectrical.cell(row=1 + 1, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 1, column=1 + 1).value = self.electrical_values["Wire diameter [mm]"]
            wElectrical.cell(row=1 + 2, column=1 + 0).value = "Wire cross sectional area [mm2]"
            wElectrical.cell(row=1 + 2, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 2, column=1 + 1).value = self.electrical_values["Wire cross sectional area [mm2]"]
            wElectrical.cell(row=1 + 3, column=1 + 0).value = "Nominal current [A]"
            wElectrical.cell(row=1 + 3, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 3, column=1 + 1).value = self.electrical_values["Nominal current [A]"]
            wElectrical.cell(row=1 + 4, column=1 + 0).value = "Maximum current [A]"
            wElectrical.cell(row=1 + 4, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 4, column=1 + 1).value = self.electrical_values["Maximum current [A]"]
            wElectrical.cell(row=1 + 5, column=1 + 0).value = "Total wire length [m]"
            wElectrical.cell(row=1 + 5, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 5, column=1 + 1).value = self.electrical_values["Total wire length [m]"]
            wElectrical.cell(row=1 + 6, column=1 + 0).value = "Wire resistance [Ohm]"
            wElectrical.cell(row=1 + 6, column=1 + 0).font = title_style
            wElectrical.cell(row=1 + 6, column=1 + 1).value = self.electrical_values["Wire resistance [Ohm]"]

            for i, coil in enumerate(self.simulation.coils):
                wCoils.cell(row=1 + i + 1, column=1 + 0).value = coil.radius
                wCoils.cell(row=1 + i + 1, column=1 + 1).value = coil.num_turns
                wCoils.cell(row=1 + i + 1, column=1 + 2).value = coil.I
                wCoils.cell(row=1 + i + 1, column=1 + 3).value = coil.pos_z

            for i, val in enumerate(self.simulation.z_arr):
                wBz.cell(row=1 + 0, column=1 + i + 1).value = val
                wBz.cell(row=1 + 0, column=1 + i + 1).font=title_style

                wBy.cell(row=1 + 0, column=1 + i + 1).value = val
                wBy.cell(row=1 + 0, column=1 + i + 1).font=title_style

                wBnorm.cell(row=1 + 0, column=1 + i + 1).value = val
                wBnorm.cell(row=1 + 0, column=1 + i + 1).font=title_style


            for i, val in enumerate(self.simulation.y_arr):
                wBz.cell(row=1 + i + 1, column=1 + 0).value = val
                wBz.cell(row=1 + i + 1, column=1 + 0).font = title_style

                wBy.cell(row=1 + i + 1, column=1 + 0).value = val
                wBy.cell(row=1 + i + 1, column=1 + 0).font = title_style

                wBnorm.cell(row=1 + i + 1, column=1 + 0).value = val
                wBnorm.cell(row=1 + i + 1, column=1 + 0).font = title_style


            for i, _ in enumerate(self.simulation.z_arr):
                for j, _ in enumerate(self.simulation.y_arr):
                    wBz.cell(row=1 + j + 1, column=1 + i + 1).value = self.simulation.Bz_grid[i, j]
                    wBy.cell(row=1 + j + 1, column=1 + i + 1).value = self.simulation.Brho_grid[i, j]
                    wBnorm.cell(row=1 + j + 1, column=1 + i + 1).value = self.simulation.norm[i, j]


            wb.save(filename)

        elif response == Gtk.ResponseType.CANCEL:
            pass
            # print("Cancel clicked")

        dialog.destroy()



    def on_import(self, widget):
        dialog = Gtk.FileChooserDialog("Please choose a file", self.window,
            Gtk.FileChooserAction.OPEN,
            (Gtk.STOCK_CANCEL, Gtk.ResponseType.CANCEL,
             Gtk.STOCK_OPEN, Gtk.ResponseType.OK))

        filters = Gtk.FileFilter()
        filters.set_name("Excel files")
        filters.add_pattern("*.*.csv")
        filters.add_pattern("*.xlsx")
        filters.add_pattern("*.XLSX")
        dialog.add_filter(filters)

        response = dialog.run()
        
        if response == Gtk.ResponseType.OK:
            filename = dialog.get_filename()


            wb = openpyxl.load_workbook(filename)
            wInput = wb["Simulation parameters"]
            wCoils = wb['Input parameters']
            wBy = wb['B y']
            wBz = wb['B z']
            wBnorm = wb['B norm']

            z_min = wInput.cell(row=1 +0, column=1 + 1).value
            z_max = wInput.cell(row=1 +1, column=1 + 1).value
            z_points = int(wInput.cell(row=1 +2, column=1 + 1).value)
            y_min = wInput.cell(row=1 +3, column=1 + 1).value
            y_max = wInput.cell(row=1 +4, column=1 + 1).value
            y_points = int(wInput.cell(row=1 +5, column=1 + 1).value)

            coils = []
            for i in range(wCoils.max_row - 1):
                radius = wCoils.cell(row=1 + i + 1, column=1 + 0).value
                turns = int(wCoils.cell(row=1 + i + 1, column=1 + 1).value)
                current = wCoils.cell(row=1 + i + 1, column=1 + 2).value
                position = wCoils.cell(row=1 + i + 1, column=1 + 3).value
                coils.append(CreateCoil("Circular", radius, turns, current, position))

            z_arr = []
            for i in range(wBz.max_column - 1):
                z_arr.append(wBz.cell(row=1 + 0, column=1 + i + 1).value)

            y_arr = []
            for i in range(wBz.max_row - 1):
                y_arr.append(wBz.cell(row=1 + i + 1, column=1 + 0).value)

            Bz_grid = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            Brho_grid = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            norm = numpy.zeros(shape=(len(z_arr), len(y_arr)))
            for i in range(len(z_arr) - 1):
                for j in range(len(y_arr) - 1):
                    Bz_grid[i, j] = wBz.cell(row=1 + j + 1, column=1 + i + 1).value
                    Brho_grid[i, j] = wBy.cell(row=1 + j + 1, column=1 + i + 1).value
                    norm[i, j] = wBnorm.cell(row=1 + j + 1, column=1 + i + 1).value

            self.simulation.set_data(coils, z_min, z_max, z_points, y_min, y_max, y_points,
                 z_arr, y_arr, Bz_grid, Brho_grid, norm)


            self.load_simulation()

        elif response == Gtk.ResponseType.CANCEL:
            # print("Cancel clicked")
            pass

        dialog.destroy()